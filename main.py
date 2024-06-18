import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter import PhotoImage
from datetime import datetime

from utils import *

import os
import sys
import requests
import shutil

from pandas import read_excel
from uuid import uuid4
import pandas as pd
from openpyxl import Workbook
from datetime import date
from config import *

# TODO(luisvasquez): To handle errors when binance API fails.


def resource_path(relative_path):
    # Get the absolute path for a resource in a PyInstaller-bundled app
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def pretty_money_format(value, scientific_threshold=1e6):
    # Format money in a human-readable way.
    if abs(value) < scientific_threshold:
        return f"{value:,.2f}"
    else:
        return f"{value:.2e}"


def get_balance_v2(wallet):
    balance = balancev2(wallet)
    return balance['data']['balance'][0]['available']


def get_balance_funding(wallet):
    balance = balancev2(wallet)
    ans = ""
    for i, item in enumerate(balance['data']['balance']):
        ans += f"{item['asset']} = {pretty_money_format(item['available'])}" + '\n'
    return ans


def download_xlsx_from_url(url, folder_path, filename):

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Create the full file path
    file_path = os.path.join(folder_path, filename)

    # Fetch the content
    response = requests.get(url, stream=True)
    response.raise_for_status()

    # Save the content to the file
    with open(file_path, 'wb') as f:
        for chunk in response.iter_content(chunk_size=8192):
            f.write(chunk)

    print(f"File downloaded as {file_path}")


class Window(tk.Tk):
    def __init__(self, title):
        super().__init__()
        self.set_theme()
        self.title(title)

    def set_theme(self):
        self.geometry("1200x600")
        self.configure(bg="#333333")
        self.option_add("*Background", "#333333")
        self.option_add("*Foreground", "#FFFFFF")
        self.option_add("*Label.Background", "#333333")
        self.option_add("*Label.Foreground", "#FFFFFF")
        self.option_add("*Entry.Background", "#555555")
        self.option_add("*Entry.Foreground", "#FFFFFF")
        self.option_add("*Button.Background", "#555555")
        self.option_add("*Button.Foreground", "#FFFFFF")

    def add_logo(self):
        self.logo_image = PhotoImage(file=resource_path("mined_logo.png"))
        self.logo_label = tk.Label(self, image=self.logo_image)
        self.logo_label.pack(padx=10, pady=10)

    def add_text_box_usr(self, text):
        self.label_username = tk.Label(self, text=text, font=FontType.FONT_14)
        self.label_username.pack(padx=10, pady=10)
        self.entry_username = tk.Entry(self, font=FontType.FONT_14)
        self.entry_username.pack()

    def add_text_box_pwd(self, text):
        self.label_password = tk.Label(self, text=text, font=FontType.FONT_14)
        self.label_password.pack(padx=10, pady=10)
        self.entry_password = tk.Entry(self, show="*", font=FontType.FONT_14)
        self.entry_password.pack()

    def add_button(self, text, command, identifier):
        setattr(self, f"{identifier}", tk.Button(
            self, text=text, command=command, font=FontType.FONT_14))
        getattr(self, f"{identifier}").pack(padx=10, pady=10)


class PayoutsWindow(Window):
    def __init__(self, title, username, role):  # TODO(luisvasquez): To add username and role
        super().__init__(title)

        top_row = tk.Frame(self)
        top_row.pack()

        middle_row = tk.Frame(self)
        middle_row.pack()

        middle_row_2 = tk.Frame(self)
        middle_row_2.pack()

        bottom_row = tk.Frame(self)
        bottom_row.pack()

        # print("pagos masivos")
        # print(username, role)

        total_amount = 0
        usdt_balance = 9999999999

        welcome_msg = tk.Label(
            top_row, text=f"Bienvenido, {username}!", font=FontType.FONT_14)
        welcome_msg.pack(padx=10, pady=10, side=tk.LEFT)

        try:
            response = balancev1("SPOT_WALLET")
            usdt_balance = response['data']['balance']
        except Exception as e:
            print(f"Failed operation when trying to get USDT balance!")

        balance_msg = tk.Label(
            top_row, text=f"Balance de la cuenta: {usdt_balance:,.2f} USDT", font=FontType.FONT_14)
        balance_msg.pack(padx=10, pady=10, side=tk.LEFT)

        try:
            data = read_excel('payments.xlsx', keep_default_na=False)
            total_amount = data['Monto'].sum()
            total_number = len(data['Monto'])
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            self.destroy()

        preview_msg1 = tk.Label(
            top_row, text=f"Número de cuentas a pagar: {total_number}", font=FontType.FONT_14)
        preview_msg1.pack(padx=10, pady=10, side=tk.LEFT)

        preview_msg2 = tk.Label(
            top_row, text=f"Total a pagar: {total_amount:.3f} USDT", font=FontType.FONT_14)
        preview_msg2.pack(padx=10, pady=10, side=tk.LEFT)

        if (total_amount > usdt_balance):
            messagebox.showerror(
                "Error", "Usted no cuenta con saldo suficiente para realizar las transacciones.")
            return

        # Prefix label and entry
        label_prefix = tk.Label(
            bottom_row, text="Prefijo (max-length=16):", font=FontType.FONT_14)
        label_prefix.pack(padx=10, pady=10, side=tk.LEFT)

        self.entry_prefix = tk.Entry(
            bottom_row, font=FontType.FONT_14, validate="key")
        self.entry_prefix.pack(side=tk.LEFT)
        self.entry_prefix.insert(0, "batch322")

        # Set the validation command to limit the maximum length
        validate_cmd = (self.register(self.validate_entry), "%P")
        self.entry_prefix.config(validate="key", validatecommand=validate_cmd)

        self.btn_update_prefix = tk.Button(
            bottom_row, text="Actualizar Prefijo", command=self.update_prefix, font=FontType.FONT_14)
        # self.btn_update_prefix.config(state=tk.DISABLED)
        self.btn_update_prefix.pack(padx=10, pady=10, side=tk.LEFT)

        self.btn_validate_payids = tk.Button(
            bottom_row, text="Validar PayIDs", command=self.validate_payids, font=FontType.FONT_14)
        # self.btn_validate_payids.config(state=tk.DISABLED)
        self.btn_validate_payids.pack(padx=10, pady=10, side=tk.LEFT)

        self.btn_confirmation = tk.Button(
            bottom_row, text="Ejecutar Pagos", command=self.show_confirmation, font=FontType.FONT_14)
        # self.btn_confirmation.config(state=tk.DISABLED)
        self.btn_confirmation.pack(padx=10, pady=10, side=tk.LEFT)

        treeview_style = ttk.Style(bottom_row)
        # Use the default theme for consistent appearance
        treeview_style.theme_use("clam")

        treeview_style.configure("Custom.Treeview",
                                 background="#333333",
                                 foreground="#FFFFFF",
                                 fieldbackground="#333333",
                                 font=FontType.FONT_14,
                                 rowheight=30)  # Increase the row height here

        # Configure a separate style for the heading cells
        treeview_style.configure("Custom.Treeview.Heading",
                                 font=FontType.FONT_14)  # Increase the font size of the headings here

        treeview_style.map("Custom.Treeview",
                           background=[("selected", "#555555")])

        # Create a treeview widget
        self.treeview = ttk.Treeview(self, columns=(
            "Name", "Amount", "Pay_ID", "Email", "IsPayIDValid", "RequestID", "Status"), style="Custom.Treeview")

        # Define column headings
        self.treeview.heading("#0", text="ID")
        self.treeview.heading("Name", text="Nombre + Descripción")
        self.treeview.heading("Amount", text="Monto (USDT)")
        self.treeview.heading("Pay_ID", text="Pay ID")
        self.treeview.heading("Email", text="Email")
        self.treeview.heading("IsPayIDValid", text="IsPayIDValid")
        self.treeview.heading("RequestID", text="RequestID")
        self.treeview.heading("Status", text="Status")

        self.bigtable = []
        for index, row in data.iterrows():
            name = row['Nombre']
            pay_id = row['Pay_ID']

            email = row['Email']
            amount = row['Monto']
            rand_id = self.entry_prefix.get() + "-" + str(uuid4()
                                                          )[0:15]  # to change it to 15

            self.treeview.insert("", "end", text=str(
                index + 1), values=(name, f"{amount:,.2f}", pay_id, email, "", rand_id, ""))
            self.bigtable.append({
                "name": name,
                "amount": amount,
                "pay_id": pay_id,
                "email": email,
                "request_id": rand_id,
                "IsPayIDValid": "",
                "status": "",
                "request_response": "",
            })

        # Format columns
        self.treeview.column("#0", width=10, anchor="center")
        self.treeview.column("Name", width=220)
        self.treeview.column("Amount", width=100, anchor="e")
        self.treeview.column("Pay_ID", width=80, anchor="center")
        self.treeview.column("Email", width=200)
        self.treeview.column("IsPayIDValid", width=100)
        self.treeview.column("RequestID", width=250)
        self.treeview.column("Status", width=50)

        # Pack the treeview widget
        self.treeview.pack(side=tk.BOTTOM, fill="both", expand=True)

    def validate_entry(self, value):
        if len(value) <= 16:
            return True
        else:
            self.entry_prefix.delete(16, tk.END)
            return False

    def update_prefix(self):
        self.treeview.delete(*self.treeview.get_children())
        for index, row in enumerate(self.bigtable):
            rand_id = self.entry_prefix.get() + "-" + str(uuid4()
                                                          )[0:15]  # change it to 15
            row['request_id'] = rand_id
            self.treeview.insert("", "end", text=str(index + 1), values=(
                row['name'], f"{row['amount']:,.2f}", row['pay_id'], row['email'], row['IsPayIDValid'], row['request_id'], row['status']))

    def validate_payids(self):
        self.treeview.delete(*self.treeview.get_children())
        for index, row in enumerate(self.bigtable):
            pay_id = row['pay_id']
            response = verification(pay_id)
            # print(response)
            row['IsPayIDValid'] = response['status']
            self.treeview.insert("", "end", text=str(index + 1), values=(
                row['name'], f"{row['amount']:,.2f}", row['pay_id'], row['email'], row['IsPayIDValid'], row['request_id'], row['status']))
        messagebox.showinfo("Mensaje", "PayIDs Validados.")

    def show_confirmation(self):
        result = messagebox.askyesno(
            "Confirmación", "Esta seguro que desea ejecutar TODOS los pagos?")
        if result:
            self.btn_update_prefix.config(state="disabled")
            self.btn_validate_payids.config(state="disabled")
            self.btn_confirmation.config(state="disabled")
            self.execute_payments()
            messagebox.showinfo("Información", "Pagos hechos.")
        else:
            pass

    def execute_payments(self):
        self.treeview.delete(*self.treeview.get_children())
        for index, row in enumerate(self.bigtable):
            amount = row['amount']
            request_id = row['request_id']
            receive_type = "PAY_ID"
            receiver = row['pay_id']
            response = payout(amount, request_id, receive_type, receiver)
            # print (response)
            row['request_response'] = response
            self.treeview.insert("", "end", text=str(index + 1), values=(
                row['name'], f"{row['amount']:,.2f}", row['pay_id'], row['email'], row['IsPayIDValid'], row['request_id'], row['status']))
        # self.iterations = 2
        self.update_status()
        self.create_logs()

    def update_status(self):
        # if self.iterations <= 0:
        #     return
        # self.iterations -= 1
        self.treeview.delete(*self.treeview.get_children())
        for index, row in enumerate(self.bigtable):
            request_id = row['request_id']
            response = query(request_id)
            row['status'] = response['status']
            self.treeview.insert("", "end", text=str(index + 1), values=(
                row['name'], f"{row['amount']:,.2f}", row['pay_id'], row['email'], row['IsPayIDValid'], row['request_id'], row['status']))
        # self.after(100, self.update_status)

    def create_logs(self):
        data = []
        field_names = ["#", "name", "ammount", "pay_id", "email", "request_id",
                       "status", "code", "error_message", "request_response", "function"]

        for index, row in enumerate(self.bigtable):
            id = row['request_id']
            name = row['name']
            amount = row['amount']
            receiver = row['pay_id']
            email = row['email']
            response = str(row['request_response'])
            status = row['request_response']['status']
            code = row['request_response']['code']
            error_message = row['request_response'].get('errorMessage')
            data.append([index + 1, name, amount, receiver, email,
                        id, status, code, error_message, response, "payout"])

        # Create a DataFrame from the data
        df = pd.DataFrame(data, columns=field_names)

        # Create a Workbook object
        workbook = Workbook()
        worksheet = workbook.active

        # Write column headers
        for col_idx, col_name in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=col_name)

        # Write data rows
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Set column widths
        column_widths = {"A": 5, "B": 30, "C": 40, "D": 10, "E": 10, "F": 70}
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Save the workbook to an Excel file
        # Set the filename based on current date
        # Set the filename based on current date and folder path
        today = date.today()
        folder = "logs"
        filename = os.path.join(
            folder, f"{today.strftime('%d-%m-%Y')}_{self.entry_prefix.get()}.xlsx")

        # Create the logs folder if it doesn't exist
        if not os.path.exists(folder):
            os.makedirs(folder)
        workbook.save(filename)


class WalletsWindow(Window):
    def __init__(self, title, username, role):
        super().__init__(title)
        # self.add_logo()

        top_row = tk.Frame(self)
        top_row.pack()

        middle_row = tk.Frame(self)
        middle_row.pack()

        middle_row_2 = tk.Frame(self)
        middle_row_2.pack()

        bottom_row = tk.Frame(self)
        bottom_row.pack()

        welcome_msg = tk.Label(
            top_row, text=f"Bienvenido, {username}!", font=FontType.FONT_14)
        welcome_msg.pack(padx=10, pady=10, side=tk.LEFT)

        balance_detail = tk.Label(
            middle_row, text=f"Funding: \n{get_balance_funding('FUNDING_WALLET')}", font=FontType.FONT_14)
        balance_detail.pack(padx=10, pady=10, side=tk.LEFT)

        balance_detail_2 = tk.Label(
            middle_row_2, text=f"Spot: \n{get_balance_funding('SPOT_WALLET')}", font=FontType.FONT_14)
        balance_detail_2.pack(padx=10, pady=10, side=tk.LEFT)

        self.add_button("Descargar Reporte",
                        self.download_report, "btn_download_01")

    def download_report(self):
        print("Report downloaded in ./reports/wallets.txt")

        dir_path = './reports'
        file_path = os.path.join(dir_path, 'wallets.txt')

        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

        with open(file_path, 'w') as f:
            print(
                f"Funding: \n{get_balance_funding(WalletType.FUNDING_WALLET)}", file=f)
            print(
                f"Spot: \n{get_balance_funding(WalletType.SPOT_WALLET)}", file=f)

        messagebox.showinfo(
            "Mensaje", "Reporte descargado en el directorio './reports/wallets'.")


class TransferWindow(Window):
    def __init__(self, title, username, role):
        super().__init__(title)

        # reemplazar por get_balance_funding('SPOT_WALLET')
        # self.spot_amount = get_balance_v2('SPOT_WALLET')
        try:
            self.spot_amount = extract_balance_v1('SPOT_WALLET')
            self.funding_amount = extract_balance_v1('FUNDING_WALLET')
        except Exception as e:
            print(f"Failed operation when trying to get USDT balance!")

        top_row = tk.Frame(self)
        top_row.pack()

        text_row = tk.Frame(self)
        text_row.pack()

        middle_row = tk.Frame(self)
        middle_row.pack()

        middle_row_2 = tk.Frame(self)
        middle_row_2.pack()

        middle_row_3 = tk.Frame(self)
        middle_row_3.pack()

        middle_row_6 = tk.Frame(self)
        middle_row_6.pack()

        middle_row_4 = tk.Frame(self)
        middle_row_4.pack()

        middle_row_5 = tk.Frame(self)
        middle_row_5.pack()

        bottom_row = tk.Frame(self)
        bottom_row.pack()

        send_button_row = tk.Frame(self)
        send_button_row.pack()

        welcome_msg = tk.Label(
            top_row, text=f"Bienvenido, {username}!", font=FontType.FONT_14)
        welcome_msg.pack(padx=10, pady=10, side=tk.LEFT)

        balance_text = tk.Label(
            text_row, text=f"Balance de cuentas", font=FontType.FONT_14)
        balance_text.pack(padx=10, pady=10, side=tk.LEFT)

        balance_detail = tk.Label(
            middle_row, text=f"Funding: \n{self.funding_amount}", font=FontType.FONT_14)
        balance_detail.pack(padx=10, pady=10, side=tk.LEFT)

        balance_detail_2 = tk.Label(
            middle_row, text=f"Spot: \n{self.spot_amount}", font=FontType.FONT_14)
        balance_detail_2.pack(padx=10, pady=10, side=tk.LEFT)
        text_amount = tk.Label(
            middle_row_2, text="Por favor seleccione el monto para hacer transferir entre cuentas", font=FontType.FONT_14)
        text_amount.pack(padx=10, pady=10)
        label_date = tk.Label(
            middle_row_3, text="Monto: ", font=FontType.FONT_14)
        label_date.pack(padx=10, pady=10, side=tk.LEFT)

        self.clear = tk.Button(
            middle_row_6, text="Limpiar", command=lambda: self.update_amount(0),  font=FontType.FONT_14)
        self.clear.pack(padx=10, pady=10, side=tk.RIGHT)

        self.btn_execute_100000 = tk.Button(
            middle_row_6, text="100 000", command=lambda: self.update_amount(100000),  font=FontType.FONT_14)
        self.btn_execute_100000.pack(padx=(400, 10), pady=10, side=tk.RIGHT)

        self.btn_execute_10000 = tk.Button(
            middle_row_3, text="10 000", command=lambda: self.update_amount(10000),  font=FontType.FONT_14)
        self.btn_execute_10000.pack(padx=10, pady=10, side=tk.RIGHT)

        self.btn_execute_1000 = tk.Button(
            middle_row_3, text="1 000", command=lambda: self.update_amount(1000), font=FontType.FONT_14)
        self.btn_execute_1000.pack(padx=10, pady=10, side=tk.RIGHT)

        self.btn_execute_100 = tk.Button(
            middle_row_3, text="100", command=lambda: self.update_amount(100),  font=FontType.FONT_14)
        self.btn_execute_100.pack(padx=10, pady=10, side=tk.RIGHT)

        self.entry_amount = tk.Entry(
            middle_row_3, font=FontType.FONT_14, validate="key", justify="right")
        self.entry_amount.pack(side=tk.LEFT)
        self.entry_amount.bind("<FocusOut>", self.on_entry_change)
        self.entry_amount.insert(0, '0')

        label_type = tk.Label(
            middle_row_3, text="USDT", font=FontType.FONT_14)
        label_type.pack(padx=10, pady=10, side=tk.LEFT)

        label_origin = tk.Label(
            middle_row_4, text="Origen: ", font=FontType.FONT_14)
        label_origin.pack(padx=10, pady=10, side=tk.LEFT)

        # destinations = ["Funding Wallet", "Spot Wallet"]

        self.combobox_destination = tk.Entry(
            middle_row_4, font=FontType.FONT_14)
        self.combobox_destination.pack(side=tk.LEFT)
        self.combobox_destination.insert(0, "Funding Wallet")
        self.combobox_destination.config(state=tk.DISABLED)

        # self.combobox_destination = ttk.Combobox(
        #     middle_row_4, values=destinations, font=FontType.FONT_14)
        # self.combobox_destination.bind(
        #     "<<ComboboxSelected>>", lambda event: self.selected_change(event))
        # self.combobox_destination.pack(side=tk.LEFT)
        # self.combobox_destination.set(destinations[0])

        label_destination = tk.Label(
            middle_row_5, text="Destino: ", font=FontType.FONT_14)
        label_destination.pack(padx=10, pady=10, side=tk.LEFT)

        self.input_origin = tk.Entry(
            middle_row_5, font=FontType.FONT_14)
        self.input_origin.pack(side=tk.LEFT)
        self.input_origin.insert(0, "Spot Wallet")
        self.input_origin.config(state=tk.DISABLED)

        self.btn_enviar = tk.Button(
            send_button_row, text="Transferir", command=self.confirmation, font=FontType.FONT_14)
        self.btn_enviar.pack(padx=10, pady=20, side=tk.LEFT)

    def on_entry_change(self, event):
        current_value = self.entry_amount.get()
        try:
            float_value = float(current_value.replace(",", ""))
            formatted_value = pretty_money_format(float_value)
            self.entry_amount.delete(0, tk.END)
            self.entry_amount.insert(0, formatted_value)
        except ValueError:
            pass

    def selected_change(self, event):
        selection = self.combobox_destination.get()
        print(f"Selección: {selection}")

        if selection == "Funding Wallet":
            self.input_origin.config(state=tk.NORMAL)
            self.input_origin.delete(0, tk.END)
            self.input_origin.insert(0, "Spot Wallet")
            self.input_origin.config(state=tk.DISABLED, fg="black")

        else:
            self.input_origin.config(state=tk.NORMAL)
            self.input_origin.delete(0, tk.END)
            self.input_origin.insert(0, "Funding Wallet")
            self.input_origin.config(state=tk.DISABLED, fg="black")
        self.input_origin.config(fg="black")

    def confirmation(self):
        amount = self.entry_amount.get()

        current_balance = ""
        if (self.combobox_destination.get() == "Spot Wallet"):
            current_balance = (self.spot_amount)
        else:
            current_balance = (self.funding_amount)

        try:
            amount_float = float(amount.replace(",", ""))
            if amount_float < 0:
                messagebox.showerror("Monto", "Ingresaste un monto negativo")
            elif amount_float > float(current_balance):
                messagebox.showerror("Monto", "Monto mayor al del origen")
                return
            else:
                message = f"¿Estás seguro de realizar la transferencia con monto {pretty_money_format(amount_float)} al destino {self.input_origin.get()}?"
                response = messagebox.askyesno("Confirmación", message)

                if response:
                    self.make_internal_transfer()
                else:
                    print("Cancelado")
        except ValueError as e:
            import traceback
            traceback.print_exc()
            print(f"Error: {e}")
            messagebox.showerror(
                "Monto", f"El valor {amount} no es un valor válido")

    def make_internal_transfer(self):
        current_date = datetime.now()
        day = current_date.date()
        time = current_date.strftime('%H:%M:%S')
        request = f"trans_{day}_{time}_prod"
        currency = "USDT"
        amount = self.entry_amount.get()
        amount = str(float(amount.replace(",", "")))
        origin = self.combobox_destination.get()

        if (origin == "Funding Wallet"):
            transfer_type = "TO_MAIN"
        else:
            transfer_type = "TO_PAY"

        response = transfer(request, currency, amount, transfer_type)

        if (response["status"] == "SUCCESS"):
            messagebox.showinfo("Completado", "Transferencia completa")
            self.btn_enviar.config(state=tk.DISABLED)
            self.btn_execute_100.config(state=tk.DISABLED)
            self.btn_execute_1000.config(state=tk.DISABLED)
            self.btn_execute_10000.config(state=tk.DISABLED)
            self.btn_execute_100000.config(state=tk.DISABLED)
            self.clear.config(state=tk.DISABLED)
            self.combobox_destination.config(state=tk.DISABLED)
            self.entry_amount.config(state=tk.DISABLED)
        else:
            messagebox.showerror("Error", "Error en transferencia")

    def update_amount(self, amount):
        if amount == 0:
            new_value = '0'
            self.entry_amount.delete(0, tk.END)
            self.entry_amount.insert(0, new_value)
            return
        current_value = self.entry_amount.get()
        try:
            current_value = float(current_value.replace(",", ""))
        except ValueError:
            current_value = 0.0

        current_balance = ""
        if (self.combobox_destination.get() == "Spot Wallet"):
            current_balance = (self.spot_amount)
        else:
            current_balance = (self.funding_amount)

        if (current_value + amount > float(current_balance)):
            return
        new_value = current_value + amount

        self.entry_amount.delete(0, tk.END)
        self.entry_amount.insert(0, pretty_money_format(new_value))


class ReportsWindow(Window):
    def __init__(self, title, username, role):
        super().__init__(title)
        # self.add_logo()

        top_row = tk.Frame(self)
        top_row.pack()

        middle_row = tk.Frame(self)
        middle_row.pack()

        middle_row_2 = tk.Frame(self)
        middle_row_2.pack()

        bottom_row = tk.Frame(self)
        bottom_row.pack()

        welcome_msg = tk.Label(
            top_row, text=f"Bienvenido, {username}!", font=FontType.FONT_14)
        welcome_msg.pack(padx=10, pady=10, side=tk.LEFT)

        # Prefix label and entry
        label_date = tk.Label(
            middle_row, text="Fecha inicial(dd/mm/yyyy):", font=FontType.FONT_14)
        label_date.pack(padx=10, pady=10, side=tk.LEFT)

        self.entry_date = tk.Entry(
            middle_row, font=FontType.FONT_14, validate="key")
        self.entry_date.pack(side=tk.LEFT)
        self.entry_date.insert(0, "01/08/2023")

        # Prefix label and entry
        label_date_2 = tk.Label(
            middle_row_2, text="Fecha Final(dd/mm/yyyy):", font=FontType.FONT_14)
        label_date_2.pack(padx=10, pady=10, side=tk.LEFT)

        self.entry_date_2 = tk.Entry(
            middle_row_2, font=FontType.FONT_14, validate="key")
        self.entry_date_2.pack(side=tk.LEFT)
        self.entry_date_2.insert(0, "15/08/2023")

        self.btn_execute = tk.Button(
            bottom_row, text="Encontrar Reportes", command=self.execute, font=FontType.FONT_14)
        self.btn_execute.pack(padx=10, pady=10, side=tk.LEFT)

    def execute(self):
        print("Updating dates...")
        print(self.entry_date.get())
        print(self.entry_date_2.get())
        response = report("Transaction", TransactionType.PAYOUT,
                          self.entry_date.get(), self.entry_date_2.get())
        print(response)

        if os.path.exists("./files"):
            shutil.rmtree("./files")

        for token in response['data']:
            link = token['downloadUrl']
            download_xlsx_from_url(link, "./files", token['fileName'])
        messagebox.showinfo(
            "Información", "Reportes descargados en la carpeta 'files'.")


class MenuWindow(Window):
    def __init__(self, title, username, role):  # TODO(luisvasquez): To add username and role
        super().__init__(title)
        self.username = username
        self.role = role
        # self.add_logo()

        self.logo_image = PhotoImage(file=resource_path("mined_logo.png"))
        self.logo_image = self.logo_image.subsample(1, 1)
        self.logo_label = tk.Label(self, image=self.logo_image)
        self.logo_label.pack(padx=10, pady=0)

        self.add_button("Pagos Masivos (Payouts)",
                        self.payouts, "btn_dashb_01")
        if role != "admin":
            self.btn_dashb_01.config(state=tk.DISABLED)
            self.btn_dashb_04.config(state=tk.DISABLED)
        self.add_button("Balance de Wallets", self.wallets, "btn_dashb_02")
        self.add_button("Reporte de Pagos", self.reports, "btn_dashb_03")
        self.add_button("Transferencia entre Cuentas",
                        self.transfers, "btn_dashb_04")
        if role != "admin":
            self.btn_dashb_03.config(state=tk.DISABLED)
        # center_and_fit(self)

    def transfers(self):
        self.destroy()
        transfer = TransferWindow(
            "Tranferencia entre Cuentas", self.username, self.role)
        transfer.attributes('-topmost', True)
        transfer.focus_force()

    def payouts(self):
        self.destroy()
        payouts = PayoutsWindow("Pagos Masivos", self.username, self.role)
        payouts.attributes('-topmost', True)  # set the window to be topmost
        payouts.focus_force()

    def wallets(self):
        self.destroy()
        wallets = WalletsWindow("Balance de Wallets", self.username, self.role)
        wallets.attributes('-topmost', True)
        wallets.focus_force()

    def reports(self):
        self.destroy()
        reports = ReportsWindow("Reporte de Pagos", self.username, self.role)
        reports.attributes('-topmost', True)
        reports.focus_force()


class LoginWindow(Window):
    def __init__(self, title):
        super().__init__(title)
        self.add_logo()
        self.add_text_box_usr(text="Usuario: ")
        self.add_text_box_pwd(text="Contraseña: ")
        self.add_button("Ingresar", self.login, "btn_login_01")

        # Additional settings
        # Set focus to the user text field
        self.entry_username.focus_set()

        # Bind Enter key press to the login button
        self.bind("<Return>", lambda event: self.btn_login_01.invoke())

        # center window
        # center_and_fit(self)

    def login(self):
        # print("Login")
        self.username = self.entry_username.get()
        self.password = self.entry_password.get()
        # print(f"Username: {self.username}")
        # print(f"Password: {self.password}")

        for credential in CREDENTIALS:
            if self.username == credential["username"] and self.password == credential["password"]:
                self.destroy()
                menu = MenuWindow("Modulos", self.username, credential["role"])
                # set the window to be topmost
                menu.attributes('-topmost', True)
                menu.focus_force()
                return

        messagebox.showerror("Login", "Usuario o contraseña incorrecto(s).")


class App(tk.Tk):
    def __init__(self, title):
        self = LoginWindow(title)

        # run
        self.mainloop()


App("Área de Tesorería - Automatizaciones App - MINED")
